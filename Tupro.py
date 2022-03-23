from openpyxl import load_workbook
import random
import xlsxwriter

def ProgramNode():
    workbook = xlsxwriter.Workbook("Node.xlsx")
    worksheet = workbook.add_worksheet('Node')
    worksheet.write('A1', 'No')
    worksheet.write('B1', 'Label')
    worksheet.write('C1', 'X')
    worksheet.write('D1', 'Y')
    
    for row in range(n):
        no = row+1
        label = no
        convert_label = 'N'+str(label) 
        x = 5* random.randint(1,20)
        y = 5* random.randint(1,20)

        worksheet.write('A'+str(no+1), no)
        worksheet.write('B'+str(no+1), convert_label)
        worksheet.write('C'+str(no+1), x)
        worksheet.write('D'+str(no+1), y)
        
        print(no, convert_label, x, y)

    workbook.close()

def Tetangga():
    workbook = xlsxwriter.Workbook("Tetangga.xlsx")
    worksheet = workbook.add_worksheet('Tetangga')
    
    huruf = 'A'

    for row in range(n):
        no_row = row+1
        convert_angka = 'N'+str(no_row)
        worksheet.write('A'+str(no_row+1), convert_angka) 
    
    for coloumn in range(n):
        no_col = coloumn+1
        huruf = chr(ord(huruf)+1)
        convert_huruf = huruf + '1'
        worksheet.write(convert_huruf, 'N'+str(no_col))

    arr = [0] * n
    next_coloumn = 0
    m = n
    next_huruf = 0
    z = 0
    for row in range(2,n+2,1):
        temp_coloumn = 0
        next_huruf2 = next_huruf
        x = 0
        y = z
        for coloumn in range(65,65+m,1):
            #print(temp_coloumn)
            #print(arr[z])
            #while arr[z] < 3:
            huruf = chr(coloumn+next_coloumn+1)
            huruf2 = chr(coloumn-next_huruf2+1)
            #print(row)
            #print(x)
            next_huruf2 += 1
            #print(coloumn+1)
            convert_huruf = huruf + str(row)
            convert_huruf2 = huruf2 + str(row+x)
            hasil_random = random.choice([0,1])
            #print(x+1,arr[x])
            if hasil_random == 1:
                temp_coloumn += 1
                if temp_coloumn <= 5 - arr[z] and arr[z]<5:
                    #print(temp_coloumn, 5-arr[x])
                    #print(x+1,arr[x])
                    worksheet.write(convert_huruf, hasil_random)
                    worksheet.write(convert_huruf2, hasil_random)
                    arr[y] += 1
                    #print("N",y+1,"=",arr[y])
                else:
                    worksheet.write(convert_huruf, 0)
                    worksheet.write(convert_huruf2, 0)
            else:
                worksheet.write(convert_huruf, hasil_random)
                worksheet.write(convert_huruf2, hasil_random)
            y += 1
            x += 1
        next_huruf -= 1
        next_coloumn += 1
        m-=1
        z+=1
    #i = 0
    #while i<n :  
    #    print("N",i+1," = ",arr[i])
     #   i+=1
    #data = load_workbook("Tetangga.xlsx")
    #sheet = data.active

    '''

    for coloumn in range(65,65+n,1):
        temp_row=0
        for row in range(2,n+2,1):
            huruf = chr(coloumn+1)
            convert_angka = huruf + str(row)
            nilai = sheet[convert_angka]
            if nilai.value == 1:
                temp_row += 1
                if temp_row == 5:
                    worksheet.write(convert_huruf, 0)

    
    for row in range(2,n+2,1):
        temp_coloumn = 0
        for coloumn in range(65,65+n,1):
            huruf = chr(coloumn+1)
            convert_huruf = huruf + str(row)
            nilai = sheet[convert_huruf]
            if nilai.value == 0:
                temp_coloumn += 1
                if temp_coloumn == 8:
                    sheet[convert_huruf] = 0
    '''
    workbook.close()


if __name__== "__main__":
    
    #Menjalankan Input n
    n = int(input("Masukkan rentang 10..20 : "))
    while (n < 10) or (n > 20):
        print("Anda Salah memasukkan input")
        n = int(input("Masukkan rentang 10..20 : "))
    
    
    ProgramNode()
    Tetangga()